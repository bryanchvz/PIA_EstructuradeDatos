import sys
import sqlite3
from sqlite3 import Error
from datetime import datetime,timedelta
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment, Border,Side
import os

def tablas():
    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("""
                              CREATE TABLE IF NOT EXISTS clientes
                              (id_cliente INTEGER PRIMARY KEY,
                              nombre TEXT NOT NULL,
                              apellidos TEXT NOT NULL)""")
            
            mi_cursor.execute("""
                              CREATE TABLE IF NOT EXISTS salas
                              (id_sala INTEGER PRIMARY KEY,
                              nombre TEXT NOT NULL,
                              cupo INTEGER NOT NULL CHECK (cupo > 0))""")
            
            mi_cursor.execute("""   
                              CREATE TABLE IF NOT EXISTS reservaciones
                              (folio INTEGER PRIMARY KEY,
                              id_cliente TEXT NOT NULL,
                              id_sala TEXT NOT NULL,
                              fecha TIMESTAMP NOT NULL,
                              turno TEXT NOT NULL,
                              evento TEXT NOT NULL,
                              estado TEXT NOT NULL DEFAULT 'ACTIVA',
                              FOREIGN KEY(id_cliente) REFERENCES clientes(id_cliente), \
                              FOREIGN KEY(id_sala) REFERENCES salas(id_sala)
                              )""")
    except Error as e:
        print(f"Error al crear las tablas {e}")

def verificar_estado_inicial():
    if os.path.exists("coworking.db"):
        print("\n=== SISTEMA DE RESERVACIÓN DE SALAS ===")
        print("Estado anterior encontrado y cargado exitosamente.\n")
    else:
        print("\n=== SISTEMA DE RESERVACIÓN DE SALAS ===")
        print("No se encontró una versión anterior del sistema.")
        print("Iniciando con estado inicial vacío...\n")
    tablas()

def registrar_cliente():
    print("\n--- REGISTRAR NUEVO CLIENTE ---")

    while True:
        nombre = input("Ingrese su nombre: ").strip()
        if nombre and nombre.replace(" ","").isalpha():
            break
        print("Error: El nombre solo puede contener letras y no puede estar vacio.")
    
    while True:
        apellidos = input("Ingrese sus apellidos: ").strip()
        if apellidos and apellidos.replace(" ","").isalpha():
            break
        print("Error: Los apellidos solo pueden contener letras y no pueden estar vacios. ")
    
    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            valores = (nombre,apellidos)
            mi_cursor.execute("INSERT INTO clientes(nombre,apellidos) \
                              VALUES(?,?)", valores)
            print(f"Cliente registrado exitosamente con clave:  {mi_cursor.lastrowid}")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

def listar_clientes():
    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT id_cliente, nombre, apellidos FROM clientes ORDER BY apellidos, nombre")
            clientes = mi_cursor.fetchall()

            if clientes:
                print("\n" + "="*70)
                print(f"{'Clave':<10} {'Apellidos':<25} {'Nombre':<25}")
                print("="*70)
                for id_cliente, nombre, apellidos in clientes:
                    print(f"{id_cliente:<10} {apellidos:<25} {nombre:<25}")
                print("="*70)
                return clientes
            else:
                print("No se encontraron clientes registrados")
                return []
    except Error as e:
        print(f"Error al listar clientes: {e}")
        return []

def registrar_sala():
    
    print("\n--- REGISTRAR NUEVA SALA ---")

    while True:
        nombre= input("Ingrese el nombre de la sala: ").strip()
        if nombre:
            break
        print("Error: El nombre de la sala no puede estar vacío.")

    while True:
        try:
            cupo = int(input("Ingrese el cupo máximo de la sala: "))
            if cupo > 0:
                break
            print("Error: El cupo debe ser mayor a 0.")
        except ValueError:
            print("Error: Debe ingresar un número válido.")

    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            valores = (nombre,cupo)
            mi_cursor.execute("INSERT INTO salas(nombre, cupo) \
                              VALUES(?,?)", valores)
            print(f"La clave asignada fue: {mi_cursor.lastrowid}")
    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

def listar_salas():
    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT id_sala, nombre, cupo FROM salas")
            salas = mi_cursor.fetchall()

            if salas:
                print("\n" + "="*70)
                print(f"{'Clave':<10} {'Nombre':<30} {'Cupo':<10}")
                print("="*70)
                for id_sala, nombre, cupo in salas:
                    print(f"{id_sala:<10} {nombre:<30} {cupo:<10}")
                print("="*70)
                return salas
            else:
                print("No hay salas registradas.")
                return []
    except Error as e:
        print(f"Error al listar salas: {e}")
        return []

def turnos_disponibles(id_sala, fecha_iso):
    todos_turnos = ["MATUTINO", "VESPERTINO", "NOCTURNO"]
    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute(
                "SELECT turno FROM reservaciones WHERE id_sala = ? AND fecha = ?",
                (id_sala, fecha_iso)
            )
            ocupados = [fila[0] for fila in mi_cursor.fetchall()]
            libres = [t for t in todos_turnos if t not in ocupados]
            return libres
        
    except Error as e:
        print(f"Error al consultar turnos disponibles: {e}")
        return []
    
def es_domingo(fecha):
    return fecha.weekday() == 6

def obtener_lunes_siguiente(fecha):
    dias_hasta_lunes = (7 - fecha.weekday()) % 7
    if dias_hasta_lunes == 0:
        dias_hasta_lunes = 1
    return fecha + timedelta(days = dias_hasta_lunes)

def registrar_reservacion():
    print("\n--- REGISTRAR RESERVACIÓN ---")

    try:
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT COUNT(*) FROM clientes")
            if mi_cursor.fetchone()[0] == 0:
                print("No hay clientes registrados.")
                return
            
            mi_cursor.execute("SELECT COUNT(*) FROM salas")
            if mi_cursor.fetchone()[0] == 0:
                print("No hay salas registradas. ")
                return
            
            listar_clientes()
            id_cliente = input("Ingrese la clave del cliente: ").strip()
            
            mi_cursor.execute("SELECT id_cliente FROM clientes WHERE id_cliente = ?", (id_cliente,))
            cliente = mi_cursor.fetchall()

            if not cliente:
                print("El cliente no existe. ")
                return
            
            listar_salas()
            id_sala = input("Ingrese la clave de la sala: ").strip()
            
            mi_cursor.execute("SELECT * FROM salas WHERE id_sala = ?", (id_sala,))

            sala = mi_cursor.fetchall()
            if not sala:
                print("No existe la sala. ")
                return
            
            fecha_str = None
            while not fecha_str:
                fecha_input = input("Ingrese la fecha de reserva (MM-DD-AAAA): ").strip()
                try:
                    fecha_dt = datetime.strptime(fecha_input, "%m-%d-%Y")

                    hoy = datetime.now()
                    diferencia_dias = (fecha_dt.date() - hoy.date()).days
                    if diferencia_dias >= 2:
                        if es_domingo(fecha_dt):
                            print("No se permiten reservaciones en domingo.")
                            lunes_siguiente = obtener_lunes_siguiente(fecha_dt)
                            
                            print(f"Se propone mover la fecha al lunes siguiente: {lunes_siguiente.strftime('%m-%d-%Y')}")
                            opcion = input("¿Desea aceptar esta nueva fecha? (s/n): ").strip().lower()
                            if opcion == "s":
                                fecha_dt = lunes_siguiente
                            else:
                                print("Ingrese una nueva fecha válida.")
                                continue

                        fecha_str = fecha_dt.strftime("%m-%d-%Y")
                    else:
                        print("Error: La fecha debe ser al menos dos días después de hoy.")
                except ValueError:
                    print("Error: Formato de fecha inválido. Use MM-DD-AAAA")

            
            fecha_iso = fecha_dt.strftime("%Y-%m-%d")

            turnos_validos = turnos_disponibles(id_sala, fecha_iso)
            if not turnos_validos:
                print(f"No hay turnos disponibles para la sala {id_sala} en la fecha {fecha_iso}.")
                return

            print(f"Turnos disponibles: {', '.join(turnos_validos)}")

            turno = None
            while not turno:
                turno_input = input("Ingrese el turno: ").strip().upper()
                if turno_input in turnos_validos:
                    turno = turno_input
                else:
                    print("Error: Turno no válido o no disponible.")

            mi_cursor.execute("SELECT * FROM reservaciones WHERE id_sala = ? AND fecha = ? AND turno = ?",
                              (id_sala,fecha_iso,turno))
            ocupado = mi_cursor.fetchall()
            if ocupado: 
                print(f"Error: La sala {id_sala} ya está reservada para el turno {turno} en esa fecha. ")
                return
            
            evento = input("Ingrese el nombre del evento: ").strip()

            mi_cursor.execute("INSERT INTO reservaciones(id_cliente, id_sala, fecha, turno, evento) \
                              VALUES (?,?,?,?,?)",
                              (id_cliente, id_sala, fecha_iso, turno, evento))
            
            print(f"Reservación registrada correctamente. Folio asignado: {mi_cursor.lastrowid}")

    except Error as e:
        print(e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

def cancelar_reservacion():
    print("\n--- CANCELAR RESERVACIÓN ---")
    try:
        fecha_inicio_input = input("Ingrese la fecha de inicio (MM-DD-AAAA): ").strip()
        fecha_inicio = datetime.strptime(fecha_inicio_input, "%m-%d-%Y").date()
        fecha_fin_input = input("Ingrese la fecha de fin (MM-DD-AAAA): ").strip()
        fecha_fin = datetime.strptime(fecha_fin_input, "%m-%d-%Y").date()

        if fecha_fin < fecha_inicio:
            print("Error: La fecha fin no puede ser menor que la fecha inicial. ")
            return
        
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("""
            SELECT r.folio, c.nombre, c.apellidos, s.nombre, r.fecha, r.turno, r.evento \
            FROM reservaciones r
            JOIN clientes c ON r.id_cliente = c.id_cliente
            JOIN salas s ON r.id_sala = s.id_sala
            WHERE r.estado = 'ACTIVA'
            AND DATE (r.fecha) BETWEEN ? AND ?
            ORDER BY r.fecha""", (fecha_inicio, fecha_fin))

            registros = mi_cursor.fetchall()
            if not registros:
                print("No hay reservaciones activas en ese rango de fechas. ")
                return
            print("\n" + "="*100)
            print(f"{'FOLIO':<8} {'CLIENTE':<25} {'SALA':<15} {'FECHA':<12} {'TURNO':<12} {'EVENTO'}")
            print("="*100)

            for folio, nombre, apellidos, sala, fecha, turno, evento in registros:
                cliente = f"{apellidos} {nombre}"
                print(f"{folio:<8} {cliente:<25} {sala:<15} {fecha:<12} {turno:<12} {evento}")
            print("="*100)

            folio_cancelar = input("Ingrese el folio de la reservacion a cancelar (o 'cancelar'): ").strip()
            if folio_cancelar.lower() == 'cancelar':
                print("Operación cancelada.")
                return
            mi_cursor.execute("SELECT fecha FROM reservaciones WHERE folio = ? AND estado = 'ACTIVA'", (folio_cancelar,))
            reserva = mi_cursor.fetchone()
            if not reserva:
                print("No se encontró una reservación activa con ese folio. ")
                return
            fecha_reserva = datetime.strptime(reserva[0], "%Y-%m-%d").date()
            if (fecha_reserva - datetime.now().date()).days < 2:
                print("Error: Solo se puede cancelar reservaciones con al menos dos días de anticipación. ")
                return
            
            confirmar = input("¿Desea confirmar la cancelación? (s/n): ").strip()
            if confirmar != 's':
                print("Cancelación abortada para el usuario. ")
                return
            mi_cursor.execute("UPDATE reservaciones SET estado = 'CANCELADA' WHERE folio = ?", (folio_cancelar,))
            conn.commit()

            print("Reservación cancelada correctamente. ")

    except ValueError:
        print("Error: Formato de fecha inválido. Use MM-DD-AAAA. ")
    except Error as e:
        print(f"Error en la cancelación: {e}")

            
def editar_evento():
    print("\n--- EDITAR NOMBRE DE EVENTO ---")

    try:    
        while True:
            try:
                fecha_inicio_input = input("Ingrese fecha de inicio (MM-DD-AAAA): ").strip()
                fecha_inicio = datetime.strptime(fecha_inicio_input, "%m-%d-%Y").date()
            except ValueError:
                print("Error: Formato de fecha inválido. Use MM-DD-AAAA.")
                continue  

            try:
                fecha_fin_input = input("Ingrese la fecha de fin (MM-DD-AAAA): ").strip()
                fecha_fin = datetime.strptime(fecha_fin_input, "%m-%d-%Y").date()
            except ValueError:
                print("Error: Formato de fecha inválido. Use MM-DD-AAAA.")
                continue

            if fecha_fin < fecha_inicio:
                print("Error: la fecha fin no puede ser menor que la fecha inicial.")
                continue
            break

        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT folio, id_cliente, id_sala, fecha, turno, evento FROM reservaciones")
            registros = mi_cursor.fetchall()

            registros_en_rango = []

            for folio, id_cliente, id_sala, fecha_str, turno, evento in registros:
                fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                if fecha_inicio <= fecha_dt <= fecha_fin:
                    registros_en_rango.append((folio, id_cliente, id_sala, fecha_dt, turno, evento))

            if not registros_en_rango:
                print("No hay reservaciones en ese rango de fechas.")
                return
            print("\n" + "="*115)
            print(f"{'FOLIO':<8} {'CLIENTE':<25} {'SALA':<20} {'FECHA':<12} {'TURNO':<12} {'EVENTO':<25}")
            print("="*115)

            for folio, id_cliente, id_sala, fecha_dt, turno, evento in registros_en_rango:
                mi_cursor.execute("SELECT nombre, apellidos FROM clientes WHERE id_cliente = ?", (id_cliente,))
                cliente = mi_cursor.fetchone()
                mi_cursor.execute("SELECT nombre FROM salas WHERE id_sala = ?", (id_sala,))
                sala = mi_cursor.fetchone()
                
                cliente_nombre = f"{cliente[1]} {cliente[0]}"
                print(f"{folio:<8} {cliente_nombre:<25} {sala[0]:<20} {fecha_dt.strftime('%m-%d-%Y'):<12} {turno:<12} {evento:<25}")

            print("="*115)


            folios_validos = {str(r[0]) for r in registros_en_rango}

            while True:
                folio_editar = input("\nIngrese el folio de la reservacion que desea editar (o 'cancelar'): ").strip()
                if folio_editar.lower() == "cancelar":
                    print("Operación cancelada.")
                    return
                if folio_editar in folios_validos:
                    break
                print("Folio inválido. Ingrese uno de los folios mostrados o escriba 'cancelar'.")

            while True:
                nuevo_evento = input("Ingrese el nuevo nombre del evento: ").strip()
                if nuevo_evento:
                    break
                print("Error: El nombre del evento no puede estar vacío.")

            mi_cursor.execute("UPDATE reservaciones SET evento = ? WHERE folio = ?", (nuevo_evento, int(folio_editar)))

            if mi_cursor.rowcount == 0:
                print("Error: No se encontró ninguna reservación con ese folio.")
            else:
                print("Evento actualizado correctamente.")

    except Error as e:
        print(e)

def exportar_reservaciones_csv(reservaciones,fecha):
    filename = f"reservaciones_{fecha.replace('-','_')}.csv"

    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Folio', 'Cliente', 'Sala', 'Cupo', 'Turno', 'Evento', 'Fecha'])

            for reserva in reservaciones:
                folio, nombre_cliente, apellidos_cliente, nombre_sala, cupo, turno, evento, fecha_reserva = reserva
                cliente_completo = f"{apellidos_cliente} {nombre_cliente}"
                writer.writerow([folio, cliente_completo, nombre_sala, cupo, turno, evento, fecha_reserva])
        
        print(f"Archivo CSV exportado correctamente: {filename}")
        return True
    except Exception as e:
        print(f"Error al exportar a CSV: {e}")
        return False

def exportar_reservaciones_json(reservaciones, fecha):
    filename = f"reservaciones_{fecha.replace('-', '_')}.json"
    
    try:
        datos_exportar = {
            'fecha_consulta': fecha,
            'total_reservaciones': len(reservaciones),
            'reservaciones': []
        }
        
        for reserva in reservaciones:
            folio, nombre_cliente, apellidos_cliente, nombre_sala, cupo, turno, evento, fecha_reserva = reserva
            datos_exportar['reservaciones'].append({
                'folio': folio,
                'cliente': {
                    'nombre': nombre_cliente,
                    'apellidos': apellidos_cliente,
                    'nombre_completo': f"{apellidos_cliente} {nombre_cliente}"
                },
                'sala': {
                    'nombre': nombre_sala,
                    'cupo': cupo
                },
                'turno': turno,
                'evento': evento,
                'fecha_reserva': fecha_reserva
            })
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(datos_exportar, f, ensure_ascii=False, indent=2)
        
        print(f"Archivo JSON exportado correctamente: {filename}")
        return True
    except Exception as e:
        print(f"Error al exportar a JSON: {e}")
        return False
    
def exportar_reservaciones_excel(reservaciones, fecha):
    filename = f"reservaciones_{fecha.replace('-', '_')}.xlsx"
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservaciones"
        
        
        ws['A1'] = f"REPORTE DE RESERVACIONES - {fecha}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        
        headers = ['Folio', 'Cliente', 'Sala', 'Cupo', 'Turno', 'Evento', 'Fecha']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thick'))
        
        
        for row_idx, reserva in enumerate(reservaciones, 3):
            folio, nombre_cliente, apellidos_cliente, nombre_sala, cupo, turno, evento, fecha_reserva = reserva
            cliente_completo = f"{apellidos_cliente} {nombre_cliente}"
            
            datos = [folio, cliente_completo, nombre_sala, cupo, turno, evento, fecha_reserva]
            for col_idx, valor in enumerate(datos, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12
        
        wb.save(filename)
        print(f"Archivo Excel exportado correctamente: {filename}")
        return True
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return False
    
def consultar_reservaciones():
    print("\n--- CONSULTAR Y EXPORTAR RESERVACIONES ---")
    
    try:
        fecha_input = input("Ingrese la fecha a consultar (MM-DD-AAAA) o Enter para hoy: ").strip()
        
        if not fecha_input:
            fecha_consulta = datetime.now()
        else:
            fecha_consulta = datetime.strptime(fecha_input, "%m-%d-%Y")
        
        fecha_formateada = fecha_consulta.strftime("%m-%d-%Y")
        fecha_iso = fecha_consulta.strftime("%Y-%m-%d")
        
        with sqlite3.connect("coworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute('''
                SELECT r.folio, c.nombre, c.apellidos, s.nombre, s.cupo, 
                       r.turno, r.evento, r.fecha
                FROM reservaciones r
                JOIN clientes c ON r.id_cliente = c.id_cliente
                JOIN salas s ON r.id_sala = s.id_sala
                WHERE r.fecha = ? AND r.estado = 'ACTIVA'
                ORDER BY r.turno, s.nombre
            ''', (fecha_iso,))
            
            reservaciones = mi_cursor.fetchall()
            
            if not reservaciones:
                print(f"No hay reservaciones para la fecha {fecha_formateada}.")
                return
            
            
            print(f"\n{'='*120}")
            print(f"RESERVACIONES PARA {fecha_formateada}".center(120))
            print(f"{'='*120}")
            print(f"{'Folio':<8} {'Cliente':<25} {'Sala':<20} {'Cupo':<8} {'Turno':<12} {'Evento':<30}")
            print(f"{'='*120}")
            
            for reserva in reservaciones:
                folio, nombre, apellidos, sala, cupo, turno, evento, fecha = reserva
                cliente = f"{apellidos} {nombre}"
                print(f"{folio:<8} {cliente:<25} {sala:<20} {cupo:<8} {turno:<12} {evento:<30}")
            
            print(f"{'='*120}")
            print(f"Total de reservaciones: {len(reservaciones)}")
            
            while True:
                print("\n--- OPCIONES DE EXPORTACIÓN ---")
                print("1. Exportar a CSV")
                print("2. Exportar a JSON")
                print("3. Exportar a Excel")
                print("4. Volver al menú principal")
                
                opcion_export = input("\nSeleccione una opción: ").strip()
                
                if opcion_export == "1":
                    exportar_reservaciones_csv(reservaciones, fecha_formateada)
                elif opcion_export == "2":
                    exportar_reservaciones_json(reservaciones, fecha_formateada)
                elif opcion_export == "3":
                    exportar_reservaciones_excel(reservaciones, fecha_formateada)
                elif opcion_export == "4":
                    break
                else:
                    print("Opción no válida. Intente nuevamente.")
                    
    except ValueError:
        print("Error: Formato de fecha inválido. Use MM-DD-AAAA.")
    except Error as e:
        print(f"Error al consultar reservaciones: {e}")
    except Exception as e:
        print(f"Error inesperado: {e}")

def menu_principal():
    verificar_estado_inicial()
    
    while True:
        print("\n" + "="*50)
        print("MENÚ PRINCIPAL".center(50))
        print("="*50)
        print("1. Registrar reservación de sala")
        print("2. Editar nombre de evento")
        print("3. Consultar reservaciones por fecha")
        print("4. Cancelar reservación")
        print("5. Registrar nuevo cliente")
        print("6. Registrar nueva sala")
        print("7. Salir")
        print("="*50)
        
        opcion = input("\nSeleccione una opción: ").strip()
        
        if opcion == '1':
            registrar_reservacion()
        elif opcion == '2':
            editar_evento()
        elif opcion == '3':
            consultar_reservaciones()
        elif opcion == '4':
            cancelar_reservacion()
        elif opcion == '5':
            registrar_cliente()
        elif opcion == '6':
            registrar_sala()
        elif opcion == '7':
            confirmacion = input("\n¿Está seguro que desea salir? (s/n): ").strip().lower()
            if confirmacion == 's':
                print("\nCerrando sistema...")
                break
        else:
            print("\nOpción no válida. Por favor intente nuevamente.")

if __name__ == "__main__":
    menu_principal()